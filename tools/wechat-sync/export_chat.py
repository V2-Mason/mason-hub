#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
微信聊天记录 BLOB 解码 + Excel 导出
======================================
前提: 你已经用 wechat-decrypt 解密了数据库
用法: python export_chat.py [decrypted目录路径]

如果不传路径, 默认读取同目录下的 decrypted/ 文件夹
"""

import os
import sys
import json
import struct
import sqlite3
import zlib
from datetime import datetime, timezone, timedelta
from pathlib import Path
from collections import defaultdict

# ============================================================
# 配置
# ============================================================
# 北京时间
BEIJING_TZ = timezone(timedelta(hours=8))

# 消息类型
MSG_TYPE = {
    1: "文本", 3: "图片", 34: "语音", 42: "名片",
    43: "视频", 47: "表情", 48: "位置", 49: "链接/文件",
    50: "通话", 10000: "系统消息", 10002: "撤回",
}


# ============================================================
# Protobuf BLOB 解码
# ============================================================
def decode_blob(blob):
    """解码微信 4.0 的 compress_content / message_content BLOB"""
    if not blob or len(blob) == 0:
        return ""

    data = blob

    # 尝试 zlib 解压
    if len(data) >= 2 and data[0] == 0x78:
        try:
            data = zlib.decompress(data)
        except:
            pass

    # 尝试 protobuf 解析 (提取 field 1 = 文本内容)
    text = extract_protobuf_string(data)
    if text:
        return text

    # 兜底: 直接 UTF-8
    try:
        t = data.decode("utf-8", errors="ignore").strip("\x00").strip()
        if t and is_readable(t):
            return t
    except:
        pass

    return ""


def extract_protobuf_string(data):
    """手动解析 protobuf, 提取 field 1 (string)"""
    offset = 0
    results = []

    while offset < len(data):
        try:
            tag, offset = read_varint(data, offset)
            if offset >= len(data):
                break

            field = tag >> 3
            wire = tag & 0x07

            if wire == 0:  # varint
                _, offset = read_varint(data, offset)
            elif wire == 1:  # 64-bit fixed
                offset += 8
            elif wire == 2:  # length-delimited (string/bytes)
                length, offset = read_varint(data, offset)
                if length < 0 or offset + length > len(data):
                    break
                value = data[offset: offset + length]
                offset += length
                if field == 1:
                    try:
                        t = value.decode("utf-8")
                        if t and is_readable(t):
                            results.append(t)
                    except:
                        pass
            elif wire == 5:  # 32-bit fixed
                offset += 4
            else:
                break
        except:
            break

    return results[0] if results else ""


def read_varint(data, offset):
    result = 0
    shift = 0
    while offset < len(data):
        b = data[offset]
        offset += 1
        result |= (b & 0x7F) << shift
        if (b & 0x80) == 0:
            return result, offset
        shift += 7
        if shift > 63:
            break
    return result, offset


def is_readable(text):
    if not text or len(text) == 0:
        return False
    ok = sum(1 for c in text if c.isprintable() or c in "\n\r\t")
    return ok / len(text) > 0.5


import re
# Excel 不支持的控制字符: 0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F
_ILLEGAL_XML_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def sanitize_for_excel(val):
    """移除 Excel/XML 不支持的控制字符"""
    if isinstance(val, str):
        return _ILLEGAL_XML_RE.sub('', val)
    return val


# ============================================================
# 读取联系人
# ============================================================
def load_contacts(dec_dir):
    """从 contact.db 读取 wxid → 昵称"""
    contacts = {}
    for db_file in dec_dir.rglob("contact*.db"):
        try:
            conn = sqlite3.connect(str(db_file))
            cur = conn.cursor()
            tables = [r[0] for r in cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()]

            for tbl in tables:
                cols = [r[1] for r in cur.execute(f"PRAGMA table_info([{tbl}])").fetchall()]
                # 微信 4.0 联系人表可能有不同列名
                name_col = next((c for c in cols if c.lower() in ("username", "usr_name", "userName")), None)
                nick_col = next((c for c in cols if c.lower() in ("nickname", "nick_name", "nickName")), None)

                if not name_col or not nick_col:
                    # 尝试 BLOB 列 (4.0 联系人也可能是 protobuf)
                    continue

                rows = cur.execute(f"SELECT [{name_col}], [{nick_col}] FROM [{tbl}]").fetchall()
                for uid, nick in rows:
                    if uid and nick:
                        contacts[uid] = nick

            conn.close()
        except:
            continue

    return contacts


# ============================================================
# 提取消息
# ============================================================
def extract_all_messages(dec_dir, contacts):
    """遍历 message_*.db, 解码 BLOB, 返回 {chat_name: [msgs]}"""
    msg_dir = dec_dir / "message"
    if not msg_dir.exists():
        # 也许 decrypted 目录结构不同, 直接搜索
        msg_files = list(dec_dir.rglob("message_[0-9]*.db"))
        if not msg_files:
            print("[错误] 找不到 message 数据库文件")
            return {}
    else:
        msg_files = sorted(msg_dir.glob("message_[0-9]*.db"))

    all_chats = defaultdict(list)
    total_msgs = 0
    decoded_ok = 0
    decoded_fail = 0

    for db_file in msg_files:
        print(f"  处理 {db_file.name} ...")
        try:
            conn = sqlite3.connect(str(db_file))
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # 读取 Name2Id (hash → wxid 映射)
            n2id = {}
            try:
                for row in cur.execute("SELECT * FROM Name2Id").fetchall():
                    d = dict(row)
                    usr = d.get("UsrName", d.get("userName", d.get("usr_name", "")))
                    if usr:
                        n2id[usr] = d
            except:
                pass

            # 找所有 Msg_ 表
            tables = [r[0] for r in cur.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name LIKE 'Msg_%'"
            ).fetchall()]

            for tbl in tables:
                tbl_hash = tbl.replace("Msg_", "")

                # 匹配联系人: 遍历 Name2Id 找对应的 wxid
                chat_id = tbl_hash
                for usr, info in n2id.items():
                    # Name2Id 通常有个列的值包含表名 hash
                    vals = [str(v) for v in info.values() if v]
                    if any(tbl_hash[:12] in v for v in vals):
                        chat_id = usr
                        break

                chat_name = contacts.get(chat_id, chat_id)
                if "@chatroom" in str(chat_id):
                    chat_name = f"[群]{chat_name}"

                try:
                    rows = cur.execute(f"SELECT * FROM [{tbl}] ORDER BY create_time ASC").fetchall()
                except:
                    try:
                        rows = cur.execute(f"SELECT * FROM [{tbl}] ORDER BY CreateTime ASC").fetchall()
                    except:
                        continue

                for row in rows:
                    d = dict(row)
                    total_msgs += 1

                    ts = d.get("create_time", d.get("CreateTime", 0))
                    is_sender = d.get("is_sender", d.get("IsSender", 0))
                    msg_type = d.get("local_type", d.get("Type", 0))

                    # === 解码消息内容 ===
                    content = ""

                    # 1) message_content (可能是明文 str 或 protobuf bytes)
                    mc = d.get("message_content")
                    if mc:
                        if isinstance(mc, str) and mc.strip():
                            content = mc.strip()
                        elif isinstance(mc, bytes) and len(mc) > 0:
                            content = decode_blob(mc)

                    # 2) compress_content (protobuf + 可能 zlib 压缩)
                    if not content:
                        cc = d.get("compress_content")
                        if cc and isinstance(cc, bytes) and len(cc) > 0:
                            content = decode_blob(cc)

                    # 3) packed_info_data
                    if not content:
                        pid = d.get("packed_info_data")
                        if pid and isinstance(pid, bytes) and len(pid) > 0:
                            content = decode_blob(pid)

                    # 4) 旧版字段
                    if not content:
                        for col in ["StrContent", "strContent", "str_content"]:
                            if col in d and d[col]:
                                content = str(d[col]).strip()
                                if content:
                                    break

                    if content:
                        decoded_ok += 1
                    else:
                        decoded_fail += 1
                        content = f"[{MSG_TYPE.get(msg_type, f'类型{msg_type}')}]"

                    # 时间
                    try:
                        dt = datetime.fromtimestamp(ts, tz=BEIJING_TZ)
                        time_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                    except:
                        time_str = str(ts)

                    all_chats[chat_name].append({
                        "time": time_str,
                        "ts": ts,
                        "sender": "我" if is_sender else chat_name,
                        "type": MSG_TYPE.get(msg_type, f"类型{msg_type}"),
                        "content": content,
                    })

            conn.close()
        except Exception as e:
            print(f"  [!!] {db_file.name} 出错: {e}")

    print(f"\n  总消息: {total_msgs}")
    print(f"  成功解码: {decoded_ok}")
    print(f"  无法解码: {decoded_fail} (图片/语音/视频等非文本消息)")

    # 排序
    for name in all_chats:
        all_chats[name].sort(key=lambda x: x["ts"])

    return dict(all_chats)


# ============================================================
# 导出 Excel
# ============================================================
def export_excel(all_chats, output_path):
    """每个联系人一个 Sheet, 第一个 Sheet 是汇总"""
    try:
        import openpyxl
    except ImportError:
        print("  安装 openpyxl ...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()

    # === 汇总 Sheet ===
    ws0 = wb.active
    ws0.title = "汇总"
    headers = ["联系人/群", "消息数", "最早消息", "最近消息"]
    ws0.append(headers)
    hfont = Font(bold=True)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    for i in range(1, 5):
        ws0.cell(1, i).font = hfont
        ws0.cell(1, i).fill = hfill

    # 按消息数排序
    sorted_chats = sorted(all_chats.items(), key=lambda x: -len(x[1]))

    sheet_count = 0
    used_names = {"汇总"}

    for chat_name, msgs in sorted_chats:
        if sheet_count >= 200:
            print(f"  [!!] Sheet 数量达到 200 上限, 剩余跳过")
            break
        if not msgs:
            continue

        # 汇总行
        ws0.append([sanitize_for_excel(chat_name), len(msgs), msgs[0]["time"], msgs[-1]["time"]])

        # 安全 Sheet 名
        sn = chat_name.replace("[群]", "群_")
        for ch in '/\\*?:[]':
            sn = sn.replace(ch, "_")
        sn = sn[:28]
        if sn in used_names:
            sn = f"{sn}_{sheet_count}"
        used_names.add(sn)

        try:
            ws = wb.create_sheet(title=sn)
        except:
            ws = wb.create_sheet(title=f"chat_{sheet_count}")

        ws.append(["时间", "发送者", "类型", "内容"])
        for i in range(1, 5):
            ws.cell(1, i).font = hfont
            ws.cell(1, i).fill = hfill

        for m in msgs:
            ws.append([
                sanitize_for_excel(m["time"]),
                sanitize_for_excel(m["sender"]),
                sanitize_for_excel(m["type"]),
                sanitize_for_excel(m["content"]),
            ])

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 60

        sheet_count += 1

    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 10
    ws0.column_dimensions["C"].width = 20
    ws0.column_dimensions["D"].width = 20

    wb.save(str(output_path))
    total_msgs = sum(len(m) for m in all_chats.values())
    print(f"\n  [OK] Excel 已保存: {output_path}")
    print(f"  [OK] {sheet_count} 个对话, {total_msgs} 条消息")


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 50)
    print("  微信聊天记录 BLOB 解码 + Excel 导出")
    print("=" * 50)

    # 确定 decrypted 目录
    if len(sys.argv) > 1:
        dec_dir = Path(sys.argv[1])
    else:
        # 默认: 同目录下的 decrypted/
        dec_dir = Path(__file__).parent / "decrypted"

    if not dec_dir.exists():
        print(f"\n[错误] 找不到 decrypted 目录: {dec_dir}")
        print(f"用法: python {Path(__file__).name} <decrypted目录路径>")
        print(f"例如: python {Path(__file__).name} C:\\Users\\hangn\\wechat-decrypt\\decrypted")
        input("\n按回车退出...")
        sys.exit(1)

    print(f"\n数据目录: {dec_dir}")

    # 读取联系人
    print("\n[1/3] 读取联系人...")
    contacts = load_contacts(dec_dir)
    print(f"  已加载 {len(contacts)} 个联系人")

    # 提取消息 + 解码 BLOB
    print("\n[2/3] 提取消息 + 解码 BLOB...")
    all_chats = extract_all_messages(dec_dir, contacts)

    if not all_chats:
        print("\n[错误] 未提取到任何消息")
        input("\n按回车退出...")
        sys.exit(1)

    # 导出 Excel
    print("\n[3/3] 导出 Excel...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(__file__).parent / f"微信聊天记录_{ts}.xlsx"
    export_excel(all_chats, out_path)

    print("\n" + "=" * 50)
    print("  完成!")
    print("=" * 50)
    input("\n按回车退出...")


if __name__ == "__main__":
    main()
